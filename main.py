"""
PDP Copy Agent — core logic layer.

Functions defined at module level here are callable from Agent Builder
blueprint Python blocks (per Writer's execution environment docs).
Blueprint blocks should stay thin: call these functions, write results
to state, and let Text Generation blocks handle LLM calls.

Pipeline:
  1. extract_product_data(file_payload) -> dict           [Python block]
  2. build_generation_prompt(product, keywords) -> str    [feeds Text Gen block]
  3. parse_copy_output(raw) -> dict                       [Python block]
  4. validate_copy(copy) -> list of violations            [Python block]
  5. build_repair_prompt(copy, violations) -> str         [feeds repair Text Gen block]
  6. hard_truncate(copy) -> dict                          [fallback if repair loop maxes out]
"""

import io
import json
import re

import writer as wf
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Constraints (from the assignment spec)
# ---------------------------------------------------------------------------

TITLE_MAX = 150
BULLET_MAX = 200
MAX_REPAIR_ATTEMPTS = 2

# Generation is a voice task -> palmyra-creative. Repair is an
# instruction-following task (shorten, preserve, return JSON) -> palmyra-x5.
GENERATION_MODEL = "palmyra-creative"
REPAIR_MODEL = "palmyra-x5"

DEFAULT_SEO_KEYWORDS = [
    "LED Desk Lamp",
    "Energy-efficient Lighting",
    "Adjustable Desk Lamp",
    "High CRI Lighting",
    "Modern Desk Lamp",
]

BRAND_VOICE = """\
Brand voice: Innovative, reliable, and customer-centric.
Tone: Professional, friendly, and approachable.
- Lead with customer benefit, then support with product capability.
- Confident but never hype-y: no exclamation marks, no "revolutionary"/"game-changing".
- Use "you/your" to speak directly to the customer.
- Weave SEO keywords in naturally; never keyword-stuff or repeat a keyword awkwardly."""


# ---------------------------------------------------------------------------
# 1. xlsx extraction
# ---------------------------------------------------------------------------

def extract_product_data(file_payload) -> dict:
    """Parse an uploaded .xlsx into a structured dict.

    Format-tolerant: handles both layouts commonly used for product info docs:
      a) key/value rows (attribute in col A, value in col B)
      b) header row + one data row per product

    `file_payload` is the Writer file-input payload: a list of dicts with
    a base64/bytes "data" field, or raw bytes when called from tests.
    """
    raw = _payload_to_bytes(file_payload)
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.worksheets[0]

    rows = [
        list(row)
        for row in ws.iter_rows(values_only=True)
        if any(c is not None and str(c).strip() for c in row)
    ]
    if not rows:
        raise ValueError("Uploaded spreadsheet contains no data.")

    # Heuristic: key/value layout if most rows have exactly 2 populated cells
    two_cell_rows = sum(
        1 for r in rows if len([c for c in r if c is not None and str(c).strip()]) == 2
    )
    if two_cell_rows >= len(rows) * 0.7:
        product = {}
        for r in rows:
            cells = [c for c in r if c is not None and str(c).strip()]
            if len(cells) >= 2:
                product[str(cells[0]).strip().rstrip(":")] = str(cells[1]).strip()
        return {"products": [product], "layout": "key_value"}

    # Otherwise: header + data rows
    headers = _dedupe_headers([
        str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])
    ])
    products = []
    for r in rows[1:]:
        products.append({
            headers[i]: str(v).strip()
            for i, v in enumerate(r)
            if i < len(headers) and v is not None and str(v).strip()
        })
    return {"products": products, "layout": "tabular"}


def _dedupe_headers(headers: list) -> list:
    """'claim', 'claim', 'claim' -> 'claim', 'claim 2', 'claim 3'.
    Real-world product docs repeat column names; without this, dict keys
    collide and all but the last duplicate silently disappear."""
    seen, out = {}, []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h} {seen[h]}")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _payload_to_bytes(file_payload) -> bytes:
    """Normalize Writer file-input payloads (or raw bytes in tests) to bytes."""
    if isinstance(file_payload, bytes):
        return file_payload
    if isinstance(file_payload, list) and file_payload:
        entry = file_payload[0]
        data = entry.get("data") if isinstance(entry, dict) else entry
        if isinstance(data, bytes):
            return data
        if isinstance(data, str):  # base64 data URL or plain base64
            import base64
            if "," in data and data.startswith("data:"):
                data = data.split(",", 1)[1]
            return base64.b64decode(data)
    raise ValueError("Unsupported file payload format from file input.")


def extract_products_from_payload(file_payload) -> dict:
    """Parse every file in an upload payload (users can upload several .xlsx
    at once, each in either layout). Pools all products found; unreadable
    files are skipped and reported rather than failing the whole batch.

    Returns {"products": [...], "files_ok": [names], "files_failed": [names]}.
    Each product dict gains a "source_file" key when the filename is known.
    """
    entries = file_payload if isinstance(file_payload, list) else [file_payload]
    products, files_ok, files_failed = [], [], []
    for i, entry in enumerate(entries):
        name = entry.get("name", f"file {i + 1}") if isinstance(entry, dict) else f"file {i + 1}"
        try:
            data = extract_product_data([entry] if isinstance(entry, dict) else entry)
        except Exception:
            files_failed.append(name)
            continue
        for p in data["products"]:
            products.append({**p, "source_file": name})
        files_ok.append(name)
    return {"products": products, "files_ok": files_ok, "files_failed": files_failed}


def product_label(product: dict, index: int = 0) -> str:
    """Human label for the product picker dropdown."""
    label = ""
    for k, v in product.items():
        if str(k).lower().strip() in ("product name", "name", "product", "title", "model"):
            label = str(v)
            break
    label = label or f"Product {index + 1}"
    src = product.get("source_file")
    return f"{label} · {src}" if src else label


# ---------------------------------------------------------------------------
# 2. Prompt builders
# ---------------------------------------------------------------------------

def build_generation_prompt(product: dict, seo_keywords: str,
                            title_max: int = TITLE_MAX,
                            bullet_max: int = BULLET_MAX) -> str:
    product_block = "\n".join(f"- {k}: {v}" for k, v in product.items())
    return f"""You are a senior e-commerce copywriter for Acme Corp.

{BRAND_VOICE}

PRODUCT INFORMATION:
{product_block}

SEO KEYWORDS (integrate naturally, highest priority first):
{seo_keywords}

Write Product Detail Page (PDP) copy with exactly this structure, returned as
JSON only — no markdown fences, no commentary:

{{
  "title": "<compelling product title, MUST be under {title_max} characters, include the primary SEO keyword>",
  "description": "<one paragraph, 80-140 words, benefit-led, weaving in 2-3 SEO keywords naturally>",
  "bullets": [
    "<5 bullet points, each MUST be under {bullet_max} characters>",
    "<each bullet: one concrete feature translated into a customer benefit>",
    "...",
    "...",
    "..."
  ]
}}

Rules:
- Only use facts present in the product information. Never invent specs.
- SEO keywords are search terms, NOT product facts. If a keyword implies a
  feature that is absent from the product information (e.g. "adjustable" when
  no adjustability is listed), you must NOT attribute that feature to the
  product. Skip that keyword entirely rather than fabricate a capability.
- Every character limit is a hard constraint. Count carefully.
- Return valid JSON and nothing else."""


def build_repair_prompt(copy: dict, violations: list,
                        title_max: int = TITLE_MAX,
                        bullet_max: int = BULLET_MAX) -> str:
    issues = "\n".join(f"- {v}" for v in violations)
    return f"""The following PDP copy violates hard character limits. Rewrite ONLY the
offending fields to fit the limits while preserving meaning, brand voice, and
SEO keywords. Leave compliant fields exactly as they are.

LIMITS: title < {title_max} chars; each bullet < {bullet_max} chars.

VIOLATIONS:
{issues}

CURRENT COPY (JSON):
{json.dumps(copy, indent=2)}

Return the full corrected copy as JSON only — same schema, no commentary."""


# ---------------------------------------------------------------------------
# 3. Output parsing
# ---------------------------------------------------------------------------

def parse_copy_output(raw: str) -> dict:
    """Parse the Text Gen block output into {title, description, bullets}.
    Tolerates markdown fences and stray prose around the JSON object."""
    text = raw.strip()
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.MULTILINE).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError(f"No JSON object found in model output: {raw[:200]}")
    copy = json.loads(match.group(0))
    for field in ("title", "description", "bullets"):
        if field not in copy:
            raise ValueError(f"Missing field '{field}' in model output.")
    if not isinstance(copy["bullets"], list):
        raise ValueError("'bullets' must be a list.")
    return copy


# ---------------------------------------------------------------------------
# 4. Validation
# ---------------------------------------------------------------------------

def validate_copy(copy: dict, title_max: int = TITLE_MAX,
                  bullet_max: int = BULLET_MAX) -> list:
    """Return a list of human-readable violations. Empty list = compliant."""
    violations = []
    title_len = len(copy.get("title", ""))
    if title_len >= title_max:
        violations.append(
            f"title is {title_len} chars (limit {title_max}): \"{copy['title']}\""
        )
    for i, b in enumerate(copy.get("bullets", []), start=1):
        if len(b) >= bullet_max:
            violations.append(
                f"bullet {i} is {len(b)} chars (limit {bullet_max}): \"{b}\""
            )
    return violations


# ---------------------------------------------------------------------------
# 5. Hard fallback (used only if repair attempts max out)
# ---------------------------------------------------------------------------

def hard_truncate(copy: dict, title_max: int = TITLE_MAX,
                  bullet_max: int = BULLET_MAX) -> dict:
    """Deterministic last resort: truncate at a word boundary. Guarantees
    compliance even if the model repeatedly fails to shorten."""
    fixed = dict(copy)
    fixed["title"] = _truncate_words(copy["title"], title_max - 1)
    fixed["bullets"] = [_truncate_words(b, bullet_max - 1) for b in copy["bullets"]]
    return fixed


def _truncate_words(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    return text[:limit].rsplit(" ", 1)[0].rstrip(",;:-")


def diff_truncated_fields(original: dict, fixed: dict) -> list:
    """Human-readable names of the fields hard_truncate actually changed.
    Feeds the human-in-the-loop review flag: truncated copy is compliant
    but may end abruptly, so a person gets the final word."""
    changed = []
    if original.get("title") != fixed.get("title"):
        changed.append("title")
    for i, (a, b) in enumerate(
        zip(original.get("bullets", []), fixed.get("bullets", [])), start=1
    ):
        if a != b:
            changed.append(f"bullet {i}")
    return changed


# ---------------------------------------------------------------------------
# 6. Presentation helper for the UI output section
# ---------------------------------------------------------------------------

def format_copy_markdown(copy: dict) -> str:
    bullets = "\n".join(f"- {b}" for b in copy["bullets"])
    return (
        f"## {copy['title']}\n\n"
        f"{copy['description']}\n\n"
        f"**Key Features**\n{bullets}"
    )


# ---------------------------------------------------------------------------
# Writer Framework state (UI bindings)
# ---------------------------------------------------------------------------

initial_state = wf.init_state({
    "seo_keywords": "",
    "seo_keywords_selected": list(DEFAULT_SEO_KEYWORDS),
    "seo_keywords_custom": "",
    "products_all": None,
    "product_options": "{}",
    "products_found": False,
    "selected_product": "",
    "title_max": str(TITLE_MAX),
    "bullet_max": str(BULLET_MAX),
    "title_chars": "",
    "longest_bullet": "",
    "bullet_count": "",
    "title_note": "",
    "bullet_note": "",
    "uploaded_file": None,
    "product_data": None,
    "gen_prompt": "",
    "repair_prompt": "",
    "violations": [],
    "pdp_copy": None,
    "pdp_copy_markdown": "",
    "needs_review": False,
    "review_note": "",
    "status_message": "Upload a product information document (.xlsx) to begin.",
})

initial_state.import_stylesheet("pdp_theme", "/static/custom.css")


# ---------------------------------------------------------------------------
# Backend-driven UI (code-managed components — versioned here, not in .wf/)
# ---------------------------------------------------------------------------
# The blueprint's UI Trigger attaches to the button via its stable id
# "pdp-generate-btn".

with wf.init_ui() as ui:
    with ui.Page({"key": "pdp", "pageMode": "compact"}, id="pdp-page"):
        ui.Heading({"headingType": "h1", "text": "PDP Copy Agent"}, id="pdp-heading")
        ui.Text(
            {"text": "Turn product specs into on-brand, SEO-aware PDP copy — "
                     "with hard character limits enforced."},
            id="pdp-subtitle",
        )
        ui.Tags(
            {
                "tags": json.dumps({
                    "acme": "Acme Corp brand voice",
                    "limits": "Hard character limits",
                    "seo": "SEO keyword aware",
                }),
                "referenceColor": "#4A46DA",
                "cssClasses": "pdp-noclick",
            },
            id="pdp-badges",
        )
        with ui.Section(
            {"title": "1 · Inputs", "containerBackgroundColor": "#fcfcfc"},
            id="pdp-inputs",
        ):
            with ui.ColumnContainer(id="pdp-columns"):
                with ui.Column({"width": "1"}, id="pdp-col-doc"):
                    ui.Heading(
                        {"headingType": "h3", "text": "Product document"},
                        id="pdp-doc-heading",
                    )
                    ui.FileInput(
                        {"label": "Product information document(s) (.xlsx)",
                         "allowFileTypes": ".xlsx",
                         "allowMultipleFiles": "yes"},
                        id="pdp-file-input",
                        binding={"wf-file-change": "uploaded_file"},
                    )
                    ui.SelectInput(
                        {"label": "Multiple products found — pick one",
                         "options": "@{product_options}",
                         "accentColor": "#4A46DA"},
                        id="pdp-product-picker",
                        binding={"wf-option-change": "selected_product"},
                        visible={"expression": "custom",
                                 "binding": "products_found",
                                 "reversed": False},
                    )
                with ui.Column({"width": "1"}, id="pdp-col-keywords"):
                    ui.Heading(
                        {"headingType": "h3", "text": "SEO keywords"},
                        id="pdp-kw-heading",
                    )
                    ui.MultiselectInput(
                        {"label": "All selected by default — unselect any that don't apply",
                         "options": json.dumps({k: k for k in DEFAULT_SEO_KEYWORDS}),
                         "accentColor": "#4A46DA"},
                        id="pdp-keywords-input",
                        binding={"wf-options-change": "seo_keywords_selected"},
                    )
                    ui.TextInput(
                        {"label": "Add custom keywords (comma-separated, optional)",
                         "placeholder": "e.g. Eye-friendly Task Light"},
                        id="pdp-keywords-custom",
                        binding={"wf-change": "seo_keywords_custom"},
                    )
            with ui.Section(
                {"title": "Advanced · Title < @{title_max} chars · "
                          "Bullets < @{bullet_max} chars each",
                 "isCollapsible": "yes",
                 "startCollapsed": "yes", "containerShadow": "none",
                 "containerBackgroundColor": "#fcfcfc"},
                id="pdp-advanced",
            ):
                with ui.ColumnContainer(id="pdp-limit-columns"):
                    with ui.Column({"width": "1"}, id="pdp-col-title-limit"):
                        ui.SelectInput(
                            {"label": "Title limit (max characters)",
                             "options": json.dumps({
                                 "100": "100 — very tight",
                                 "150": "150 — standard (default)",
                                 "200": "200 — marketplace long-form",
                             }),
                             "accentColor": "#4A46DA"},
                            id="pdp-title-limit",
                            binding={"wf-option-change": "title_max"},
                        )
                    with ui.Column({"width": "1"}, id="pdp-col-bullet-limit"):
                        ui.SelectInput(
                            {"label": "Per-bullet limit (max characters)",
                             "options": json.dumps({
                                 "150": "150 — very tight",
                                 "200": "200 — standard (default)",
                                 "250": "250 — detailed",
                             }),
                             "accentColor": "#4A46DA"},
                            id="pdp-bullet-limit",
                            binding={"wf-option-change": "bullet_max"},
                        )
            ui.Button(
                {"text": "Generate PDP Copy", "lucideIcon": "sparkles",
                 "buttonColor": "#4A46DA", "buttonTextColor": "#ffffff"},
                id="pdp-generate-btn",
            )
        ui.Message({"message": "@{status_message}"}, id="pdp-status")
        with ui.Section(
            {"title": "2 · PDP Copy", "containerBackgroundColor": "#F5F5F9",
             "containerShadow": "none"},
            id="pdp-output",
        ):
            with ui.ColumnContainer(
                id="pdp-metrics-cols",
                visible={"expression": "custom", "binding": "pdp_copy_markdown",
                         "reversed": False},
            ):
                with ui.Column({"width": "1"}, id="pdp-metric-col-title"):
                    ui.Metric(
                        {"name": "Title length", "metricValue": "@{title_chars} chars",
                         "note": "@{title_note}"},
                        id="pdp-metric-title",
                    )
                with ui.Column({"width": "1"}, id="pdp-metric-col-bullet"):
                    ui.Metric(
                        {"name": "Longest bullet", "metricValue": "@{longest_bullet} chars",
                         "note": "@{bullet_note}"},
                        id="pdp-metric-bullet",
                    )
                with ui.Column({"width": "1"}, id="pdp-metric-col-count"):
                    ui.Metric(
                        {"name": "Bullet points", "metricValue": "@{bullet_count}",
                         "note": "+ benefit-led"},
                        id="pdp-metric-count",
                    )
            ui.Text(
                {"text": "@{review_note}", "cssClasses": "pdp-review-flag"},
                id="pdp-review-flag",
                visible={"expression": "custom", "binding": "needs_review",
                         "reversed": False},
            )
            ui.Text(
                {"text": "@{pdp_copy_markdown}", "useMarkdown": "yes",
                 "quickCopy": "yes"},
                id="pdp-copy-text",
                visible={"expression": "custom", "binding": "pdp_copy_markdown",
                         "reversed": False},
            )
            ui.Text(
                {"text": "Your generated copy will appear here."},
                id="pdp-copy-placeholder",
                visible={"expression": "custom", "binding": "pdp_copy_markdown",
                         "reversed": True},
            )
